import openpyxl, argparse, csv, sys, os
from openpyxl.utils import get_column_letter
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

#Directory is being changed to make it easier
os.chdir('/tmp/otrs_reports')

#parser is created
parser = argparse.ArgumentParser(description='Adds .csv to existing .xlsx',
	usage='''How to use:
			\nOnly add the lines from week 20:
			\notrs_reporting.py -w 20''')

parser.add_argument('-w', type=int, help='Number of the week')

args = parser.parse_args()

#check if args.w is there and set weekVar accordingly
if args.w:
	weekVar = args.w
else:
	sys.exit()

lastWeek = weekVar - 1
csvFileName = 'answers_' + str(weekVar) + '.2018.csv'
oldFileName = 'Antworten_' + str(lastWeek) + '.2018.xlsx'

for folder, subfolder, files in os.walk(os.getcwd()):
	for file in files:
		if file == oldFileName:
			wb = openpyxl.load_workbook(file) #open the Excel file
			sheet = wb.active #get the sheet

			max_column = sheet.max_column #get the max column
			max_row = sheet.max_row #get the max row
			

			csvFile = open(csvFileName, encoding='ISO-8859-1') #create the new .csv file and set the encoding
			csvReader = csv.reader(csvFile)
			headers = next(csvReader) #remove the first line of the .csv

			csv_row = 0 #a counter needed for a max row variable

			for row in csvReader: #loop through the .csv file
				if row[5] == str(weekVar): #check if the row is the same as the week provided
					csv_row += 1 #the counter gets increased for every row

			new_max_row = max_row + csv_row #set the new max, which is the size of the .xlsx after the .csv entries have been added
			csvFile.close() #can't loop through a csv twice

			#same as above; open a .csv and remove the first line
			csvFile = open(csvFileName, encoding='ISO-8859-1')
			csvReader = csv.reader(csvFile)
			headers = next(csvReader)

			for cellObj,csvRow in zip(sheet['A'+str(max_row+1):'F'+str(new_max_row)],csvReader): #loop through the new max rows of the .xlsx sheet and the reader (both have to have the same size and they should)
				for cell,counter in zip(cellObj,range(6)): #loop through the cell object to get the cells and a counter variable to get the .csv row indexes
					if counter == 1: #check if the current field is a ticket_id
						hyperlink = '=HYPERLINK("https://otrs.tis-cca.com/otrs/index.pl?Action=AgentTicketZoom;TicketID=' + str(csvRow[1]) + '",' + str(csvRow[1]) + ')' #assemble the hyperlink
						csvRow[counter] = hyperlink #the current field in the list is being replaced by the hyperlink
					cell.value = csvRow[counter] #write the value to the cell
					
					if counter == 1: #check again if it's a ticket_id
						cell.font = Font(color=colors.BLUE, underline="single") #set some styling

			csvFile.close() #close the original .csv
			newXLSXName = 'Antworten_' + str(weekVar) + '.2018.xlsx'
			wb.save(newXLSXName) #save the new .xlsx

