import csv, openpyxl, os
from openpyxl.utils import get_column_letter
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

#Directory is being changed to make it easier
os.chdir('/tmp/otrs_reports')

#Get all folders and files
for folder, subfolder, files in os.walk(os.getcwd()):
	for file in files:
		#Files are looped through and only .csv files are interesting
		if file.endswith('.csv'):
			#New name is being created for later
			newFileName = os.path.splitext(file)[0] + '.xlsx'
			
			#The file is opened with a specific encoding
			queryFile = open(file, encoding='ISO-8859-1')
			exampleReader = csv.reader(queryFile)
			
			#The Excelsheet is being created
			wb = openpyxl.Workbook()
			sheet = wb.active
			
			#A counter for getting the Excel cells
			i = 1

			#Loop through the .csv Object
			for row in exampleReader:
				#Another counter
				count = 1
				while count <= len(row):
					#If the current field in the .csv is a ticket_id
					if (count-2) == 0 and i != 1:
						#The hyperlink formulas is being created and being styled
						hyperlink = '=HYPERLINK("https://otrs.com/otrs/index.pl?Action=AgentTicketZoom;TicketID=' + str(row[count-2]) + '",' + str(row[count-2]) + ')'
						sheet[get_column_letter(count) + str(i)].value = hyperlink
						sheet[get_column_letter(count) + str(i)].font = Font(color=colors.BLUE, underline="single")
					#If it's not a ticket_id
					else:
						sheet[get_column_letter(count) + str(i)].value = row[count-2]
					count += 1
				i += 1

#			Delete last row
#			for cellObj in sheet['A'+str(sheet.max_row):str(get_column_letter(sheet.max_column))+str(sheet.max_row)]:
#				for cell in cellObj:
#					cell.value = ''

			#Filter the sheet
			sheet_range = 'A1' + ':' + str(get_column_letter(sheet.max_column)) + str(sheet.max_row)
			sheet.auto_filter.ref = sheet_range
			sheet.auto_filter.add_filter_column(0,[])
			
			#Save the Excel sheet
			wb.save(newFileName)
			queryFile.close()
