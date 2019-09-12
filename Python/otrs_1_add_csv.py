import openpyxl, argparse, csv, sys, os
from openpyxl.utils import get_column_letter

Directory is being changed to make it easier
os.chdir('/tmp/otrs_reports')

#parser is created
parser = argparse.ArgumentParser(description='Adds .csv to existing .xlsx and convert it all to .csv',
	usage='''How to use:
			\nOnly add the lines from week 20:
			\notrs_reporting.py -w 20
			\nAdd all entries:
			\notrs_reporting.py -a all''')

parser.add_argument('-w', type=str, help='Number of the week')
parser.add_argument('-a', type=str, help='all for all entries')

args = parser.parse_args()

#check if args.w is there and set weekVar and also set/disable allVar
if args.w:
	weekVar = args.w
	allVar = 0
	
else:
	#opposite of above. weekVar is set/disabled and allVar is set
	weekVar = 0
	if args.a:
		allVar = 1
	#if nothing gets provided exit with an help message
	else:
		print('Please enter a week with -w or select all entries with -a all')
		sys.exit()



wb = openpyxl.load_workbook('Antworten_22.2018.xlsx') #open the Excel file
sheet = wb.active #get the sheet

max_column = sheet.max_column #get the max column

#csvName = file.rstrip('.xlsx') #create the base for the .csv by stripping the .xlsx
csvName = 'Antworten_23.2018'
csvOutputFile = open(csvName + '.csv', 'w', newline='', encoding='ISO-8859-1') #create the csv file
csvOutputWriter = csv.writer(csvOutputFile) #create the writer using the file

for i in range(1,sheet.max_row): #increase the counter up to the max row, starting from 1
	for cellObj in sheet['A'+str(i):get_column_letter(max_column)+str(i)]: #loop through the Excel
		cellList = [] #set the list with the cells to nothing, because it is being appended to the csv later
		for cell in cellObj: #loop through the current row
			if type(cell.value) is float: #check if the value is a float
				cell.value = round(cell.value,2) #and round it if it is (otherwise you get ugly numbers)
			cellList.append(cell.value) #append the current value to the list of cells
		csvOutputWriter.writerow(cellList) #write the full row to the .csv
		
wb.close() #the .xlsx is no longer needed


csvFile = open('answers_23.2018.csv', encoding='ISO-8859-1') #create the new .csv file and set the encoding
exampleReader = csv.reader(csvFile)
headers = next(exampleReader) #remove the first line of the .csv

#loop through the rows
for row in exampleReader:
	new_row = [] #the new row is always being changed, so it needs to emptied
	if row[4] == weekVar: #check if the provided week matches the row's week
		hyperlink = '=HYPERLINK("https://otrs.com/otrs/index.pl?Action=AgentTicketZoom;TicketID=' + str(row[0]) + '",' + str(row[0]) + ')' #assemble the hyperlink
		new_row.extend((row[5],hyperlink,row[1],row[2],row[3],row[4])) #create the contents for the index
		csvOutputWriter.writerow(new_row) #write the contents to the .csv
		
	elif allVar == 1: #check if all entries should be added
		#same process as above
		hyperlink = '=HYPERLINK("https://otrs.com/otrs/index.pl?Action=AgentTicketZoom;TicketID=' + str(row[0]) + '",' + str(row[0]) + ')'
		new_row.extend((row[5],hyperlink,row[1],row[2],row[3],row[4]))
		csvOutputWriter.writerow(new_row)
	
	
csvFile.close() #close the original .csv
csvOutputFile.close() #close the new .csv



